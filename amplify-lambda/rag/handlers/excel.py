# Copyright (c) 2024 Vanderbilt University
# Authors: Jules White, Allen Karns, Karely Rodriguez, Max Moundas

import openpyxl
import io
from enum import Enum

from rag.handlers.text import TextExtractionHandler
from rag.handlers.shared_functions import (
    hash_visual_data,
    ensure_supported_format,
    format_visual_chunk_data,
)


PNG = "image/png"


class VisualType(Enum):
    """Enum for visual content types we extract from Excel"""

    IMAGE = "Image"  # Photos, screenshots, logos, imported images


def wrap_comma_with_quotes(s):
    if "," in s:
        s = '"' + s + '"'
    return s


class ExcelHandler(TextExtractionHandler):

    def extract_text(self, file_content, visual_map={}):
        """
        Extract text and visual content from Excel file.
        Now supports visual_map from preprocessing for multimodal content.
        """
        with io.BytesIO(file_content) as f:
            workbook = openpyxl.load_workbook(f, data_only=True)

            # Preprocess visual_map to group visuals by sheet name for efficient lookup
            visuals_by_sheet = {}
            for visual_marker, visual_data in visual_map.items():
                sheet_name = visual_data.get("location", {}).get("sheet_name")
                if sheet_name:
                    if sheet_name not in visuals_by_sheet:
                        visuals_by_sheet[sheet_name] = []
                    visuals_by_sheet[sheet_name].append(visual_data)

            chunks = []
            for sheet_number, sheet_name in enumerate(workbook.sheetnames, start=1):
                current_sheet = workbook[sheet_name]

                # Extract tabular data (rows)
                for row_number, row in enumerate(
                    current_sheet.iter_rows(values_only=True), start=1
                ):
                    row_text = ",".join(
                        wrap_comma_with_quotes(str(cell))
                        for cell in row
                        if cell is not None
                    )
                    row_text = row_text.strip()

                    if row_text:  # Only add non-empty rows
                        chunks.append(
                            {
                                "content": row_text,
                                "tokens": self.num_tokens_from_string(row_text),
                                "location": {
                                    "sheet_number": sheet_number,
                                    "sheet_name": current_sheet.title,
                                    "row_number": row_number,
                                },
                                "canSplit": False,
                            }
                        )

                # Process any visuals that belong to this sheet (O(1) lookup)
                sheet_visuals = visuals_by_sheet.get(sheet_name, [])
                for visual_data in sheet_visuals:
                    # Check if visual has been processed (has transcription)
                    if visual_data.get("transcription"):
                        # Create visual chunk using the format function
                        visual_chunk = format_visual_chunk_data(visual_data, self.num_tokens_from_string)
                        chunks.append(visual_chunk)

            return chunks

    ### Visual Data Extraction ###
    def preprocess_excel_visuals(self, file_content):
        """
        Extract visual content and create a visual map for LLM processing.
        Modifies the Excel file to include visual markers that MarkItDown will preserve.
        """

        # First pass: Extract visual data with images
        visual_map = {}
        with io.BytesIO(file_content) as f:
            workbook_with_images = openpyxl.load_workbook(f, data_only=False)
            
            # Process each sheet to extract visuals
            for sheet_idx, sheet_name in enumerate(workbook_with_images.sheetnames, start=1):
                sheet = workbook_with_images[sheet_name]
                sheet_visuals = self.extract_sheet_visuals(sheet, sheet_idx, sheet_name)
                visual_map.update(sheet_visuals)

        # Second pass: Create clean workbook without images and inject markers
        with io.BytesIO(file_content) as f:
            clean_workbook = openpyxl.load_workbook(f, data_only=True)  # No images
            
            # Inject markers into clean workbook
            for sheet_idx, sheet_name in enumerate(clean_workbook.sheetnames, start=1):
                sheet = clean_workbook[sheet_name]
                
                # Find markers for this sheet
                sheet_markers = {k: v for k, v in visual_map.items() 
                               if v['location']['sheet_name'] == sheet_name}
                
                # Inject markers
                for marker, visual_data in sheet_markers.items():
                    self._inject_visual_marker_into_sheet(sheet, marker, visual_data)

        # Convert clean workbook to bytes
        modified_content = self._workbook_to_bytes(clean_workbook)
        return modified_content, visual_map

    def _inject_visual_marker_into_sheet(self, sheet, marker, visual_data):
        """
        Inject a visual marker into the Excel sheet at an appropriate location
        """
        cell_range = visual_data['location']['cell_range']
        
        try:
            # Try to parse the cell range to find insertion point
            if ':' in cell_range:
                # Range like "A1:C3" - use the top-left cell
                start_cell = cell_range.split(':')[0]
            else:
                # Single cell like "A1"
                start_cell = cell_range
            
            # Find an appropriate cell to insert the marker
            target_cell = self._find_insertion_cell(sheet, start_cell)
            
            if target_cell:
                # Insert the marker text
                sheet[target_cell] = marker
                # print(f"Injected {marker} into cell {target_cell}")
            else:
                # Fallback: add to the end of the sheet
                self._append_marker_to_sheet(sheet, marker)
                
        except Exception as e:
            print(f"Error injecting marker {marker}: {e}")
            # Fallback: add to the end of the sheet
            self._append_marker_to_sheet(sheet, marker)

    def _find_insertion_cell(self, sheet, reference_cell):
        """
        Find an appropriate cell to insert the visual marker near the reference cell
        """
        try:
            # Convert cell reference to row/column
            cell_obj = sheet[reference_cell]
            row = cell_obj.row
            col = cell_obj.column
            
            # Strategy: Look for empty cells near the image location
            # Check cells in this order: right, below, left, above
            candidates = [
                (row, col + 1),      # Right
                (row + 1, col),      # Below  
                (row, col - 1),      # Left (if col > 1)
                (row - 1, col),      # Above (if row > 1)
            ]
            
            for candidate_row, candidate_col in candidates:
                if candidate_row > 0 and candidate_col > 0:
                    candidate_cell = sheet.cell(candidate_row, candidate_col)
                    if candidate_cell.value is None:
                        return f"{openpyxl.utils.get_column_letter(candidate_col)}{candidate_row}"
            
            # If no empty adjacent cells, try to find empty cell in same row
            for offset in range(1, 10):  # Check next 10 columns
                check_col = col + offset
                if check_col <= sheet.max_column + 5:  # Don't go too far right
                    candidate_cell = sheet.cell(row, check_col)
                    if candidate_cell.value is None:
                        return f"{openpyxl.utils.get_column_letter(check_col)}{row}"
            
            return None
            
        except Exception as e:
            print(f"Error finding insertion cell: {e}")
            return None

    def _append_marker_to_sheet(self, sheet, marker):
        """
        Append visual marker to the end of the sheet as fallback
        """
        try:
            # Find the last row with data
            last_row = sheet.max_row
            
            # Add marker in column A, a few rows below the data
            target_row = last_row + 2
            sheet[f"A{target_row}"] = marker
            # print(f"Appended {marker} to row {target_row}")
            
        except Exception as e:
            print(f"Error appending marker: {e}")

    def _workbook_to_bytes(self, workbook):
        """
        Convert workbook back to bytes for MarkItDown processing
        """
        # Convert to bytes for MarkItDown
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def extract_sheet_visuals(self, sheet, sheet_number, sheet_name):
        """Extract all visual content from a single Excel sheet"""

        visuals = {}
        # Count each visual type we extract
        type_counters = {
            VisualType.IMAGE.value: 0,
        }

        # Extract Images
        if hasattr(sheet, "_images") and sheet._images:
            for image in sheet._images:
                type_counters[VisualType.IMAGE.value] += 1
                marker = f"<Visual#{sheet_number}_{VisualType.IMAGE.value}_{type_counters[VisualType.IMAGE.value]}>"

                visual_data = self.extract_image_data(image, sheet_number, sheet_name)
                if visual_data:
                    visuals[marker] = visual_data

        return visuals

    def extract_image_data(self, image, sheet_number, sheet_name):
        """Extract image data from Excel sheet"""
        print(f"Extracting image data from sheet {sheet_name}")

        try:
            # Get image bytes
            if hasattr(image, "_data"):
                image_bytes = image._data()
            elif hasattr(image, "ref"):
                # Handle image reference
                image_bytes = image.ref
            else:
                print(f"Could not extract image data from sheet {sheet_name}")
                return None

            # Determine original format
            original_format = getattr(image, "format", "image/png")

            # Convert to supported format if necessary
            final_image_bytes, final_format = ensure_supported_format(
                image_bytes, original_format
            )

            # Generate hash for deduplication
            content_hash = hash_visual_data(final_image_bytes)

            # Get image position
            cell_range = self.get_image_cell_range(image)

            # Extract metadata
            metadata = self.extract_image_metadata(image)

            return {
                "type": VisualType.IMAGE.value,
                "format": final_format,
                "data": final_image_bytes,
                "hash": content_hash,
                "location": {
                    "sheet_number": sheet_number,
                    "sheet_name": sheet_name,
                    "cell_range": cell_range,
                },
                "alt_text": metadata.get("alt_text", ""),
                "title": metadata.get("title", ""),
                "hyperlink": metadata.get("hyperlink", ""),
                "original_format": original_format,
            }

        except Exception as e:
            print(f"Image extraction failed on sheet {sheet_name}: {e}")
            return None

    ### Helper Methods ###

    def get_image_cell_range(self, image):
        """Get the cell range where image is positioned"""
        try:
            if hasattr(image, "anchor"):
                return self.anchor_to_cell_range(image.anchor)
        except:
            pass
        return "Unknown"

    def anchor_to_cell_range(self, anchor):
        """Convert anchor to cell range notation"""
        try:
            if hasattr(anchor, "_from"):
                from_col = anchor._from.col
                from_row = anchor._from.row
                from_cell = (
                    f"{openpyxl.utils.get_column_letter(from_col + 1)}{from_row + 1}"
                )

                if hasattr(anchor, "to"):
                    to_col = anchor.to.col
                    to_row = anchor.to.row
                    to_cell = (
                        f"{openpyxl.utils.get_column_letter(to_col + 1)}{to_row + 1}"
                    )
                    return f"{from_cell}:{to_cell}"
                else:
                    return from_cell
        except:
            pass
        return "Unknown"

    def extract_image_metadata(self, image):
        """Extract metadata from image"""
        metadata = {}

        try:
            metadata["title"] = getattr(image, "name", "")
            metadata["alt_text"] = getattr(image, "description", "")
            metadata["hyperlink"] = ""

            # Try to get hyperlink if present
            if hasattr(image, "hyperlink"):
                metadata["hyperlink"] = str(image.hyperlink)

        except Exception as e:
            print(f"Error extracting image metadata: {e}")

        return metadata
