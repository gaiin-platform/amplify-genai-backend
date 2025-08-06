# Copyright (c) 2024 Vanderbilt University
# Authors: Jules White, Allen Karns, Karely Rodriguez, Max Moundas

import openpyxl
import io

from rag.handlers.text import TextExtractionHandler


class ExcelHandler(TextExtractionHandler):
    def extract_text(self, file_content, encoding):
        with io.BytesIO(file_content) as f:
            # Use openpyxl to load the workbook
            workbook = openpyxl.load_workbook(
                f, data_only=True
            )  # data_only=True to read values only, not formulas

            chunks = []
            for sheet_number, sheet in enumerate(workbook.sheetnames, start=1):
                current_sheet = workbook[sheet]

                for row_number, row in enumerate(
                    current_sheet.iter_rows(values_only=True), start=1
                ):
                    row_text = " ".join(str(cell) for cell in row if cell is not None)
                    row_text = row_text.strip()

                    chunks.append(
                        {
                            "content": row_text,
                            "location": {
                                "sheet_number": sheet_number,
                                "sheet_name": current_sheet.title,
                                "row_number": row_number,
                            },
                            "canSplit": False,
                        }
                    )

            return chunks

    def handle(self, file_content, key, split_params):
        with io.BytesIO(file_content) as f:
            # Use openpyxl to load the workbook
            workbook = openpyxl.load_workbook(
                f, data_only=True
            )  # data_only=True to read values only, not formulas

            min_chunk_size = split_params.get("min_chunk_size", 512)

            chunks = []
            for sheet_number, sheet in enumerate(workbook.sheetnames, start=1):
                current_sheet = workbook[sheet]
                current_chunk_content = ""
                current_chunk_location = None

                for row_number, row in enumerate(
                    current_sheet.iter_rows(values_only=True), start=1
                ):
                    row_text = " ".join(str(cell) for cell in row if cell is not None)
                    row_text = row_text.strip()

                    if (
                        row_text
                        and (len(current_chunk_content) + len(row_text))
                        >= min_chunk_size
                    ):
                        if (
                            current_chunk_content
                        ):  # If there is existing text, include it as a new chunk
                            chunks.append(
                                {
                                    "content": current_chunk_content,
                                    "location": current_chunk_location,
                                }
                            )
                            current_chunk_content = ""
                        current_chunk_location = {
                            "sheet_number": sheet_number,
                            "sheet_name": current_sheet.title,
                            "row_number": row_number,
                        }
                    current_chunk_content += (
                        " " + row_text if current_chunk_content else row_text
                    )

                # If there is remaining text after the loop, add it as the last chunk
                if current_chunk_content:
                    chunks.append(
                        {
                            "content": current_chunk_content,
                            "location": current_chunk_location,
                        }
                    )

            return chunks
